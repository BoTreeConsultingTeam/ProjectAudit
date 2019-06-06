from django.shortcuts import render, redirect,get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from .audit_forms import Process_Checklist_Entry, AuditForm, ChecklistForm, MixForm
from django.views import View,generic
from .models import Process_Checklist, Audits, Audit_Items
from project.models import Project
from user.models import CustomUser
from django.contrib.auth.mixins import *
from django.views.generic import UpdateView,DeleteView,TemplateView
from django.shortcuts import render, redirect,render_to_response,get_object_or_404
from datetime import date, datetime
from django.db import IntegrityError, transaction
import json
import jsonpickle
from .tasks import *
from .models import *
import numpy as nmp
import pylab as pyl
from django.db.models import Q
from django.utils.dateparse import parse_date
from rest_framework.views import APIView
from rest_framework.response import Response
from collections import OrderedDict
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, Color
from decimal import Decimal
#import Image, StringIO



class AuditExcelSave(generic.DetailView):
	def get(self, request, **kwargs):
		wb = Workbook()
		ws = wb.active
		ws.title = "Audits Summary"
		ws.sheet_properties.tabColor = "1072BA"
		print(ws)
		projectid=self.kwargs['pk']
		projectname=Project.objects.get(id=projectid)
		users = Audits.objects.filter(project=projectid).order_by('-id')
		ft = Font(bold=True)
		ws['A1']="Title"
		ws['B1']="Date"
		ws['C1']="Status"
		ws['D1']="Average Rating"
		ws['E1']="General Comment"
		ws['A1'].font = ft
		ws['B1'].font = ft
		ws['C1'].font = ft
		ws['D1'].font = ft
		ws['E1'].font = ft
		row=2
		ar=0
		for user in users:
			ws['A'+str(row)]=user.audit_title
			ws['B'+str(row)]=str(user.date)
			ws['C'+str(row)]=user.status
			ws['D'+str(row)]=user.average_rating
			ar+=user.average_rating
			ws['E'+str(row)]=user.general_comment
			row+=1
		ar=ar/len(users)
		print(type(ar))
		ar=round(ar,2)
		print(ar)
		ws['E'+str(row+1)]="Average Rate = "+str(ar)
		ws['E'+str(row+1)].font = ft
		for user in users:
			ws1 = wb.create_sheet(str(user.audit_title))
			id1 = user.id
			x = Audit_Items.objects.filter(audit_id=id1)
			ws1['A1']="Checklist Title"
			ws1['B1']="Comment"
			ws1['C1']="Rating"
			ws1['A1'].font = ft
			ws1['B1'].font = ft
			ws1['C1'].font = ft
			row=2
			for y in x:
				ws1['A'+str(row)]=y.checklist.checklist_title
				ws1['B'+str(row)]=y.comment
				ws1['C'+str(row)]=y.rating
				row+=1
			ws1['C'+str(row)]="Average Rating = "+str(user.average_rating)
		wb.save('Project - '+str(projectname)+'.xlsx')
		wb = load_workbook('Project - '+str(projectname)+'.xlsx')
		print(wb)
		return redirect('auditreportview')









class ChecklistEntry(View):
	def get(self, request):
		form = Process_Checklist_Entry()
		return render(request, 'audit/checklistentry.html', {'form': form})
	def post(self, request):
		if request.method == 'POST':
		    form = Process_Checklist_Entry(request.POST)
		    if form.is_valid():
		        print("hi")
		        form.save()
		        return redirect('checklistview')
		else:
		    form = Process_Checklist_Entry()
		return render(request, 'audit/checklistentry.html', {'form': form})

class ChecklistView(LoginRequiredMixin,generic.ListView):
	login_url = '/accounts/login/'
	template_name = 'audit/checklistview.html'
	context_object_name = 'user_list'
	
	def get_queryset(self):
		#print(Process_Checklist.objects.all())
		return Process_Checklist.objects.all()

class ChecklistUpdate(LoginRequiredMixin,UpdateView):
	model = Process_Checklist
	form_class = Process_Checklist_Entry
	template_name = 'audit/checklistupdate.html'
	success_url = '/audit/checklistview'

class ChecklistDelete(LoginRequiredMixin,DeleteView):
	model = Process_Checklist
	template_name = 'audit/checklistdelete.html'
	success_url = '/audit/checklistview'


class AuditReportCreate(View):
	def get(self, request):
		form = AuditForm()
		sub_form = ChecklistForm(prefix="check")
		checklist=Process_Checklist.objects.all()
		group=Process_Checklist.objects.all().values('group').distinct()
		#print(form)
		return render(request, 'audit/auditreport.html', {'form': form, 'checklist': checklist, 'group': group, 'sub_form': sub_form})
	def post(self, request):
		if request.method == 'POST':
			try:
				with transaction.atomic():
					form = AuditForm(request.POST)
					sub_form = ChecklistForm(request.POST, prefix="check")
					#print(sub_form)
					sub_form.audit_id=1
					#sub_form.checklist_id=
					rate=0
					items=Process_Checklist.objects.all()
					if form.is_valid():
						#print("hi")
						form.average_rating=rate/len(items)
						#print(form,form.average_rating)
						#print("hi")
						form.save()
						#return redirect('checklistview')

					audits=Audits.objects.all()
					auditid=audits[len(Audits.objects.all())-1].id
					#auditid=1
					count=0
					for item in items:
						sub_form.comment=request.POST.get('comment-'+str(item.id))
						sub_form.rating=request.POST.get('rating-'+str(item.id))
						#print("rating= ", sub_form.rating)
						if sub_form.rating != "0":
							count+=1
						rate+=int(request.POST.get('rating-'+str(item.id)))
						sub_form.date=date.today()
						sub_form.audit_id=auditid
						sub_form.checklist_id=item.id
						#print(sub_form.comment, sub_form.rating, sub_form.date, sub_form.audit_id, sub_form.checklist_id)
						audit_items=Audit_Items.objects.create(comment=sub_form.comment, rating=sub_form.rating, date=date.today(), audit_id=auditid, checklist_id=item.id)
						audit_items.save()
						'''print(sub_form)
						if sub_form.is_valid():
							print("asdasidl")
							print(sub_form)
							sub_form.save()
							#return redirect('checklistview')'''
					#print(count)
					Audits.objects.filter(id=auditid).update(average_rating=rate/count)
					sender={}
					sender['comment']=request.POST['general_comment']
					sender['rating']=str(Audits.objects.get(id=auditid).average_rating)
					sender['admin_email']=str(CustomUser.objects.get(designation="Admin").email)
					sender['lead_email']=str(Project.objects.get(id=request.POST['project']).team_lead.email)
					sender['admin_username']=str(CustomUser.objects.get(designation="Admin").username)
					sender['lead_username']=str(Project.objects.get(id=request.POST['project']).team_lead.username)
					#print(sender['lead_email'])
					sender['title']=request.POST['audit_title']
					sender['date']=request.POST['date']
					mails.delay(sender)
					return redirect('checklistview')
			except IntegrityError:
				print("One of the trasnsaction didn't work properly so all transactions rolled back")


		else:
			form = AuditForm()
		return render(request, 'audit/auditreport.html', {'form': form})


class AuditReportView(LoginRequiredMixin,generic.ListView):
	login_url = '/accounts/login/'
	template_name = 'audit/auditreportview.html'
	context_object_name = 'user_list'
	
	def get_queryset(self):
		#print(Audits.objects.all())
		users=Audits.objects.all()
		return {'users':users}

class AuditUpdate(LoginRequiredMixin,UpdateView):
	#model = Audits
	#form_class = MixForm
	#fields = ['project', 'audit_title', 'general_comment', 'status','date','comment','rating']
	#template_name = 'audit/auditupdate.html'
	#success_url = '/audit/auditreportview'

	'''def get_context_data(self, **kwargs):
		context = super(AuditUpdate, self).get_context_data(**kwargs)
		print(type(self.kwargs['pk']))
		context['Audit_Items'] = Audit_Items.objects.filter(audit_id=self.kwargs['pk']) #whatever you would like
		print(context)
		return context'''

	def get(self, request, **kwargs):

		audit=Audits.objects.filter(id=self.kwargs['pk'])
		audit_items=Audit_Items.objects.filter(audit_id=self.kwargs['pk'])
		checklist=Process_Checklist.objects.all()
		project=Project.objects.all()
		#print(audit)
		#print("hi \n")
		#print(audit_items)
		return render(request,'audit/auditupdate.html',{'audit':audit,'audit_items':audit_items, 'checklist':checklist, 'project':project})	
		return self.render_to_response(self.get_context_data())

	def post(self, request, **kwargs):
		if request.method == 'POST':
			try:
				with transaction.atomic():
					form = AuditForm(request.POST)
					sub_form = ChecklistForm(request.POST, prefix="check")
					#print(sub_form)
					sub_form.audit_id=1
					#sub_form.checklist_id=
					#print(self.kwargs['pk'])
					rate=0
					items=Process_Checklist.objects.all()
					#print(form)
					if form.is_valid():
						#print("hi")
						form.average_rating=rate/len(items)
						#print(form,form.average_rating)
						#print("hi")
						#print(form.audit_title)
						#print(datetime.strptime(str(request.POST.get('date')), "%Y-%m-%d").date())
						#date = parse_date(str(request.POST.get('date')))
						n=str(request.POST['date']).split('/')
						#print(n)
						#print(date(int(n[2]),int(n[1]),int(n[0])))
						print(n)
						Audits.objects.filter(id=self.kwargs['pk']).update(project=request.POST['project'],audit_title=request.POST['audit_title'],date=date(int(n[2]),int(n[1]),int(n[0])),status=request.POST['status'],general_comment=request.POST['general_comment'],average_rating=form.average_rating)

						#return redirect('checklistview')

					audits=Audits.objects.all()
					auditid=self.kwargs['pk']
					#auditid=1
					count=0
					for item in items:
						sub_form.comment=request.POST.get('comment-'+str(item.id))
						sub_form.rating=request.POST.get('rating-'+str(item.id))
						#print("rating = ", sub_form.rating, type(sub_form.rating))
						if sub_form.rating != "0":
							count+=1
						#print(request.POST['comment-'+str(item.id)])
						#print(request.POST['rating-'+str(item.id)])
						rate+=int(request.POST.get('rating-'+str(item.id)))
						sub_form.date=date.today()
						sub_form.audit_id=self.kwargs['pk']
						sub_form.checklist_id=item.id
						#print(sub_form.comment, sub_form.rating, sub_form.date, sub_form.audit_id, sub_form.checklist_id)
						audit_items=Audit_Items.objects.filter(Q(checklist=item.id) & Q(audit=self.kwargs['pk'])).update(comment=sub_form.comment, rating=sub_form.rating, date=date.today(), audit_id=auditid, checklist_id=item.id)
						#audit_items.save()
						'''print(sub_form)
						if sub_form.is_valid():
							print("asdasidl")
							print(sub_form)
							sub_form.save()
							#return redirect('checklistview')'''
					#print(count)
					Audits.objects.filter(id=auditid).update(average_rating=rate/count)
					return redirect('checklistview')
			except IntegrityError:
				print("One of the trasnsaction didn't work properly so all transactions rolled back")


		else:
			form = AuditForm()
		return render(request, 'audit/auditreport.html', {'form': form})	


class AuditDelete(LoginRequiredMixin, DeleteView):
	
	def get(self, request, **kwargs):
		audit=Audits.objects.get(id=self.kwargs['pk'])
		return render(request,'audit/auditreportdelete.html',{'audit':audit})	

	def post(self, request, **kwargs):
		if request.method == 'POST':
			try:
				with transaction.atomic():
					a=Audits.objects.filter(id=self.kwargs['pk']).delete()
					print(a)
					b=Audit_Items.objects.filter(audit=self.kwargs['pk']).delete()
					print(b)
					return redirect('auditreportview')
			except IntegrityError:
				print("One of the trasnsaction didn't work properly so all transactions rolled back")



class AuditItemsView(View):
	
	model = Audit_Items
	template_name = 'audit/audititemsview.html'
	'''def get(self, request, *args, **kwargs):
		self.object = self.get_object()
		id = kwargs.get('pk')
		print(id)
		context = self.get_context_data(object=self.object)
		print(context)
		return render(request,'audit/audititemsview.html')'''

	def get(self, request):
		id1 = request.GET.get('id')
		x = Audit_Items.objects.filter(audit_id=id1)
		audit=Audits.objects.get(id=id1)
		#print(x)
		return render(request, 'audit/audititemsview.html', {'x': x, 'audit': audit})

class AuditChart(generic.DetailView):
	def get(self, request, *args, **kwargs):
		x=list(Audits.objects.all().values_list('date',flat=True))
		y=list(Audits.objects.all().values_list('average_rating',flat=True))
		projectid=self.kwargs['pk']
		projectname=Project.objects.get(id=projectid)
		users = Audits.objects.filter(project=projectid).order_by('-id')
		#print(users)
		#context = {}
		#context['python_object'] = jsonpickle.encode(user)
		#print(x)
		#print(y)

		#return render('audit/auditchart.html', context)
		data=zip(x,y)
		return render(request, 'audit/auditchart.html', {'x': x, 'y': y, 'data':data, 'users':users, 'projectid':projectid, 'projectname':projectname})
	template_name = 'audit/auditchart.html'
	context_object_name = 'user_list'

	def get_queryset(self):
		#print(Audits.objects.all())
		xy=list(Audits.objects.filter(project=5).values('date','average_rating'))
		x = list(Audits.objects.filter(project=5).values_list('date', flat=True))
		y = list(Audits.objects.filter(project=5).values_list('average_rating', flat=True))
		#print(x,y)
		std1=[5,6,7,8]
		spi1=[7.5,8.5,8.2,9.0]
		std2=[5,6,7,8]
		spi2=[6.2,7.2,8.8,9.5]
		fig = pyl.figure(figsize=(5,3),dpi=200)
		#fig.set_size_inches(9.35,3.00)
		pyl.plot(x,y,linestyle="solid",marker="o",color='red')
		#pyl.plot(std2,spi2,linestyle="solid",marker="o",color='green')
		pyl.title('Project Data')
		pyl.xlabel('Date')
		pyl.ylabel('Rating')
		#pyl.xlim(0.0,10.0)
		#pyl.ylim(0.0,10.0)
		#pyl.show()
		pyl.savefig('static/test.png')

		#df=pd.DataFrame({'x': range(1,10), 'y': np.random.randn(9)*80+range(1,10) })
		return xy

def get_data(request, *args, **kwargs):
	x=list(Audits.objects.all().values_list('date',flat=True))
	y=list(Audits.objects.all().values_list('average_rating',flat=True))
	user = Audits.objects.all()
	#print(x)
	#print(y)
	data=zip(x,y)
	return JsonResponse(data)

class ListGraph(APIView):
	def get(self, request, format=None, *args, **kwargs):
		#print(self.kwargs['pk'])
		labels=list(Audits.objects.filter(project=self.kwargs['pk']).values_list('date',flat=True))
		default_items=list(Audits.objects.filter(project=self.kwargs['pk']).values_list('average_rating',flat=True))
		user = Audits.objects.all()
		data={"labels": labels, "default": default_items}
		return Response(data)

class PieGraph(APIView):
	def get(self, request, format=None, *args, **kwargs):
		a=Audits.objects.order_by('project__id','-date').distinct('project__id').values_list('status', flat=True)
		data1=OrderedDict()
		data1['Blue']=0
		data1['Orange']=0
		data1['Red']=0
		data1['Green']=0

		for i in a:
			if i == "Blue":
				data1['Blue']+=1
			elif i == "Orange":
				data1['Orange']+=1
			elif i == "Red":
				data1['Red']+=1
			elif i == "Green":
				data1['Green']+=1
		#print(list(data1.values()))
		data={"labels": list(data1.keys()), "default": list(data1.values())}

		a1=Audits.objects.order_by('project__id','-date').distinct('project__id').values_list('average_rating', flat=True)
		data2={'1 Rating':0,'2 Rating':0,'3 Rating':0,'4 Rating':0,'5 Rating':0,'6 Rating':0,'7 Rating':0,'8 Rating':0, '9 Rating':0, '10 Rating':0}
		for i in a1:
			if i == 1:
				data2['1 Rating']+=1
			elif i == 2:
				data2['2 Rating']+=1
			elif i == 3:
				data2['3 Rating']+=1
			elif i == 4:
				data2['4 Rating']+=1
			elif i == 5:
				data2['5 Rating']+=1
			elif i == 6:
				data2['6 Rating']+=1
			elif i == 7:
				data2['7 Rating']+=1
			elif i == 8:
				data2['8 Rating']+=1
			elif i == 9:
				data2['9 Rating']+=1
			elif i == 10:
				data2['10 Rating']+=1
		#print(list(data1.values()))
		data={"labels": list(data1.keys()), "default": list(data1.values()), "labels1": list(data2.keys()), "default1": list(data2.values())}
		return Response(data)


'''def CityLookup(request,pk):
    cou=serializers.serialize('json',City.objects.all().filter(region_id=pk))
    cou1=serializers.serialize('json',City.objects.all().filter(region_id=pk))
    j={'cou':cou, 'cou1':cou1}

    return JsonResponse(j,safe=False)'''
	
# Create your views here.
